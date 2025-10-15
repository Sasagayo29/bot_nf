import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import threading
import os
import shutil
import re
from datetime import datetime
import openpyxl
import fitz


class App(tk.Tk):
    """
    Classe principal da aplicação com interface gráfica (GUI).
    """

    def __init__(self):
        super().__init__()

        self.title("Automatizador de Padronização de NFs e Boletos")
        self.geometry("800x600")

        # Variáveis para armazenar os caminhos
        self.excel_path = tk.StringVar()
        self.pdf_folder_path = tk.StringVar()

        # --- Layout da Interface ---
        main_frame = ttk.Frame(self, padding="10")
        main_frame.pack(fill="both", expand=True)

        # Seção de seleção de arquivos
        selection_frame = ttk.LabelFrame(
            main_frame, text="1. Selecione os Arquivos e Pastas", padding="10")
        selection_frame.pack(fill="x", pady=5)
        selection_frame.columnconfigure(1, weight=1)

        # Selecionar arquivo Excel
        ttk.Label(selection_frame, text="Planilha de Controle:").grid(
            row=0, column=0, padx=5, pady=5, sticky="w")
        excel_entry = ttk.Entry(
            selection_frame, textvariable=self.excel_path, state="readonly", width=80)
        excel_entry.grid(row=0, column=1, padx=5, pady=5, sticky="ew")
        ttk.Button(selection_frame, text="Selecionar...", command=self.select_excel_file).grid(
            row=0, column=2, padx=5, pady=5)

        # Selecionar pasta de PDFs
        ttk.Label(selection_frame, text="Pasta de NFs/Boletos:").grid(row=1,
                                                                      column=0, padx=5, pady=5, sticky="w")
        pdf_folder_entry = ttk.Entry(
            selection_frame, textvariable=self.pdf_folder_path, state="readonly", width=80)
        pdf_folder_entry.grid(row=1, column=1, padx=5, pady=5, sticky="ew")
        ttk.Button(selection_frame, text="Selecionar...", command=self.select_pdf_folder).grid(
            row=1, column=2, padx=5, pady=5)

        # Seção de execução
        run_frame = ttk.Frame(main_frame, padding="10")
        run_frame.pack(fill="x", pady=10)
        self.run_button = ttk.Button(
            run_frame, text="2. Iniciar Processamento", command=self.start_processing_thread)
        self.run_button.pack()

        self.progress_bar = ttk.Progressbar(
            run_frame, orient="horizontal", mode="determinate")
        self.progress_bar.pack(fill="x", pady=10)

        # Seção de Log
        log_frame = ttk.LabelFrame(
            main_frame, text="Log de Atividades", padding="10")
        log_frame.pack(fill="both", expand=True, pady=5)
        self.log_area = scrolledtext.ScrolledText(
            log_frame, wrap=tk.WORD, state="disabled")
        self.log_area.pack(fill="both", expand=True)

    def select_excel_file(self):
        path = filedialog.askopenfilename(
            title="Selecione a planilha de controle",
            filetypes=(("Arquivos Excel", "*.xlsx;*.xls"),
                       ("Todos os arquivos", "*.*"))
        )
        if path:
            self.excel_path.set(path)
            self.log_message(f"Planilha selecionada: {path}")

    def select_pdf_folder(self):
        path = filedialog.askdirectory(
            title="Selecione a pasta com as subpastas 'Notas' e 'Boletos'")
        if path:
            self.pdf_folder_path.set(path)
            self.log_message(f"Pasta de PDFs selecionada: {path}")

    def log_message(self, message):
        """Adiciona uma mensagem na área de log da interface."""
        self.log_area.config(state="normal")
        self.log_area.insert(
            tk.END, f"{datetime.now().strftime('%H:%M:%S')} - {message}\n")
        self.log_area.see(tk.END)  # Auto-scroll
        self.log_area.config(state="disabled")
        self.update_idletasks()  # Força a atualização da interface

    def update_progress(self, value):
        self.progress_bar['value'] = value
        self.update_idletasks()

    def start_processing_thread(self):
        """Inicia o processamento em uma thread separada para não travar a GUI."""
        excel = self.excel_path.get()
        pdf_folder = self.pdf_folder_path.get()

        if not excel or not pdf_folder:
            messagebox.showerror(
                "Erro", "Por favor, selecione a planilha e a pasta de PDFs antes de iniciar.")
            return

        self.run_button.config(state="disabled")
        self.log_area.config(state="normal")
        self.log_area.delete('1.0', tk.END)
        self.log_area.config(state="disabled")

        # A função de processamento será executada em outra thread
        thread = threading.Thread(
            target=process_files,
            args=(excel, pdf_folder, self.log_message,
                  self.update_progress, self.processing_finished)
        )
        thread.daemon = True
        thread.start()

    def processing_finished(self, success, message):
        """Callback executado ao final do processamento."""
        self.run_button.config(state="normal")
        if success:
            messagebox.showinfo("Processo Concluído", message)
        else:
            messagebox.showerror("Erro no Processamento", message)


def extract_invoice_number_from_pdf(pdf_path):
    """
    Abre um PDF, extrai todo o texto e tenta encontrar o número da nota fiscal
    usando uma série de padrões de expressão regular (regex).
    """
    try:
        with fitz.open(pdf_path) as doc:
            full_text = ""
            for page in doc:
                full_text += page.get_text()

        if not full_text:
            return None, ""

        # O VBA original usa a primeira linha para decidir qual padrão usar.
        # Vamos simplificar e tentar uma lista de padrões robustos em ordem de prioridade.
        # Esses padrões são a tradução direta dos padrões mais eficazes do VBA.
        patterns = [
            re.compile(r"NOTA FISCAL INDICADA AO LADO\s+([\d.]+)"),
            re.compile(
                r"NF-e\s*N[ºo\.]*\s*[:\.\-]*\s*([0-9]{9}|[0-9]{3}\.[0-9]{3}\.[0-9]{3})", re.IGNORECASE),
            re.compile(
                r"SÉRIE[:\s\r\n]+(\d{3,9}(?:[.\s]?\d{1,3}){0,2})", re.IGNORECASE),
            re.compile(
                r"N[ºro\.]?[r]?\s*[:\.\s]*([0-9]{5,9}|[0-9]{1,3}\.[0-9]{3}\.[0-9]{3}|[0-9]{2,3}\.[0-9]{3})", re.IGNORECASE),
            re.compile(
                r"NOTA\s+FISCAL\s+ELETRÔNICA[\s\S]{0,50}?([0-9]{3}\.?[0-9]{3}\.?[0-9]{3})", re.IGNORECASE),
        ]

        for pattern in patterns:
            match = pattern.search(full_text)
            if match:
                numero_nota = match.group(1).strip()
                return numero_nota, full_text

        return None, full_text  # Retorna o texto mesmo que o número não seja encontrado
    except Exception as e:
        return None, f"Erro ao ler PDF: {e}"


def process_files(excel_path, pdf_folder_path, log_callback, progress_callback, finish_callback):
    """
    Função principal que executa toda a lógica de negócio.
    """
    try:
        log_callback("Iniciando processamento...")
        log_callback(
            "Carregando planilha Excel. Isso pode levar um momento...")

        # --- 1. Ler e analisar a planilha do Excel com openpyxl ---
        # Usar openpyxl permite lidar com células mescladas, assim como o VBA fazia.
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.worksheets[1]  # Segunda planilha

        # Encontra o cabeçalho
        header_row_idx = -1
        for i in range(1, 10):  # Procura até a linha 10
            if str(ws.cell(row=i, column=2).value).strip().upper() == "TRANSMISSORAS":
                header_row_idx = i
                break

        if header_row_idx == -1:
            raise ValueError(
                "Não foi possível encontrar o cabeçalho 'TRANSMISSORAS' na coluna B.")

        # Encontra as colunas de interesse
        col_indices = {}
        for i in range(1, 50):  # Procura até a coluna 50
            header = str(ws.cell(row=header_row_idx,
                         column=i).value).strip().upper()
            if header == "PEDIDO":
                col_indices['pedido'] = i
            elif header == "NF":
                col_indices['nf'] = i
            elif header in ["DANFE", "Nº DANFE"]:
                col_indices['danfe'] = i

        if 'pedido' not in col_indices or 'nf' not in col_indices:
            raise ValueError(
                "Colunas 'PEDIDO' ou 'NF' não encontradas na planilha.")

        log_callback("Analisando dados da planilha...")
        # Dicionário para mapear a chave de referência (coluna A) aos dados do pedido
        # Formato: {chave_ref: [pedido, transmissora, contador, linha]}
        pedidos_data = {}

        # Loop principal para extrair dados, tratando células mescladas
        # Esta lógica replica o comportamento do VBA de agrupar por pedido
        for row in range(header_row_idx + 1, ws.max_row + 1):
            pedido_cell = ws.cell(row=row, column=col_indices['pedido'])

            # Pula linhas vazias
            if pedido_cell.value is None:
                continue

            # Obtem o valor da célula mesclada
            for merged_range in ws.merged_cells.ranges:
                if pedido_cell.coordinate in merged_range:
                    pedido_cell = ws.cell(
                        row=merged_range.min_row, column=merged_range.min_col)
                    break

            pedido_valor = str(pedido_cell.value)

            # Lógica para agrupar NFs dentro de um Pedido
            # Simplificação: Em vez de um dicionário complexo, vamos mapear cada NF individualmente
            nf_cell = ws.cell(row=row, column=col_indices['nf'])
            if nf_cell.value is not None:
                chave_ref = str(ws.cell(row=row, column=1).value).strip()
                if chave_ref:
                    transmissora = str(
                        ws.cell(row=row, column=2).value).strip()
                    # Remove parênteses e conteúdo extra da transmissora
                    transmissora = re.sub(r'\(.*\)', '', transmissora).strip()

                    # O VBA tinha uma lógica complexa para um contador. Vamos simplificar
                    # assumindo que cada linha de NF é única.
                    if chave_ref not in pedidos_data:
                        pedidos_data[int(chave_ref)] = [
                            pedido_valor, transmissora, 0, row]

        log_callback(
            f"{len(pedidos_data)} registros de NF encontrados na planilha.")

        # --- 2. Criar pasta de destino ---
        parent_folder = os.path.dirname(pdf_folder_path)
        current_month = datetime.now()
        # Mapeia número do mês para nome em português
        month_names = ["", "Janeiro", "Fevereiro", "Março", "Abril", "Maio",
                       "Junho", "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"]
        folder_name = f"{current_month.strftime('%m')} - {month_names[current_month.month]}"
        output_folder = os.path.join(parent_folder, folder_name)
        os.makedirs(output_folder, exist_ok=True)
        log_callback(f"Pasta de destino criada/verificada: {output_folder}")

        # --- 3. Processar arquivos PDF ---
        files_to_process = []
        for root, _, files in os.walk(pdf_folder_path):
            folder_name = os.path.basename(root).upper()
            if folder_name in ["NOTAS", "BOLETOS"]:
                for filename in files:
                    if filename.lower().endswith('.pdf'):
                        files_to_process.append(
                            (os.path.join(root, filename), folder_name))

        total_files = len(files_to_process)
        progress_callback(0)

        processed_count = 0
        for pdf_path, subfolder_name in files_to_process:
            filename = os.path.basename(pdf_path)
            log_callback(
                f"Processando '{filename}' na pasta '{subfolder_name}'...")

            # Extrai a chave numérica do nome do arquivo
            match = re.match(r'^\s*(\d+)', filename)
            if not match:
                log_callback(
                    f"AVISO: Não foi possível extrair a chave numérica de '{filename}'. Pulando.")
                continue

            file_key = int(match.group(1))

            if file_key in pedidos_data:
                pedido_info = pedidos_data[file_key]
                pedido_num = pedido_info[0]
                transmissora = pedido_info[1]
                # contador = pedido_info[2] # A lógica de contador do VBA era complexa. simplificando.
                row_to_update = pedido_info[3]

                tipo_arq = ""
                col_to_update = 0

                if subfolder_name == "BOLETOS":
                    tipo_arq = "Boleto"
                    col_to_update = 28  # Coluna AB
                else:  # NOTAS
                    tipo_arq = "Danfe"
                    col_to_update = 29  # Coluna AC

                # Verifica se já foi processado
                status_cell = ws.cell(row=row_to_update, column=col_to_update)
                if status_cell.value and "Processado" in str(status_cell.value):
                    log_callback(
                        f"AVISO: '{filename}' já foi processado anteriormente. Pulando.")
                    continue

                new_filename = f"{pedido_num} - {transmissora} - {tipo_arq}.pdf"
                new_filepath = os.path.join(output_folder, new_filename)

                # Trata arquivos com nomes duplicados
                counter = 1
                while os.path.exists(new_filepath):
                    new_filename = f"{pedido_num} - {transmissora} - {tipo_arq} ({counter}).pdf"
                    new_filepath = os.path.join(output_folder, new_filename)
                    counter += 1

                shutil.copy2(pdf_path, new_filepath)
                log_callback(
                    f"  -> Arquivo copiado e renomeado para: {new_filename}")

                # Se for uma NOTA, extrai o número e atualiza a planilha
                if subfolder_name == "NOTAS":
                    numero_nota, texto_completo = extract_invoice_number_from_pdf(
                        pdf_path)

                    if numero_nota:
                        log_callback(
                            f"  -> Número da nota encontrado: {numero_nota}")
                        # Formata o número da nota como no VBA
                        numero_nota_limpo = re.sub(r'[^\d]', '', numero_nota)
                        numero_nota_formatado = numero_nota_limpo.zfill(9)
                        ws.cell(row=row_to_update, column=col_indices.get(
                            'danfe', 27)).value = numero_nota_formatado
                    else:
                        log_callback(
                            f"  -> AVISO: Número da nota não encontrado em '{filename}'.")

                    # Atualiza o texto completo da nota
                    ws.cell(row=row_to_update,
                            column=27).value = texto_completo  # Coluna AA

                # Marca como processado
                ws.cell(row=row_to_update,
                        column=col_to_update).value = f"{tipo_arq} Processado"
            else:
                log_callback(
                    f"AVISO: Chave '{file_key}' do arquivo '{filename}' não encontrada na planilha.")

            processed_count += 1
            progress_callback((processed_count / total_files) * 100)

        log_callback("Salvando alterações na planilha Excel...")
        wb.save(excel_path)
        log_callback("Planilha salva com sucesso.")

        final_message = f"Processo finalizado com sucesso!\n\nArquivos foram salvos em:\n{output_folder}"
        finish_callback(True, final_message)

    except Exception as e:
        error_message = f"Ocorreu um erro inesperado: {e}"
        log_callback(f"ERRO: {error_message}")
        import traceback
        log_callback(traceback.format_exc())
        finish_callback(False, error_message)


if __name__ == "__main__":
    app = App()
    app.mainloop()
